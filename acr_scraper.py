"""
ACR Convergence 2026 Conference Scraper
Target: https://acrconvergence2026.eventscribe.net/agenda.asp?pfp=Browse%20by%20Day
Filter: Full Schedule (All Days)

Features:
1. Multi-threaded AJAX fetching (SessionInfo & PresentationInfo popups)
2. On-disk JSON/HTML caching for fast idempotent re-runs
3. Comprehensive extraction of Session & Presentation details, Faculty/Presenters, Affiliations, Disclosures
4. Master Excel export (acr_2026_programme.xlsx) with professional OpenPyXL styling
"""

from __future__ import annotations

import html as _html
import json
import logging
import os
import re
import sys
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Optional
from urllib.parse import urljoin

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.StreamHandler(sys.stdout)],
)
if hasattr(sys.stdout, "reconfigure"):
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass
log = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Configuration & Constants
# ---------------------------------------------------------------------------

BASE_URL = "https://acrconvergence2026.eventscribe.net"
AGENDA_ALL_URL = f"{BASE_URL}/agenda.asp?BCFO=&pfp=Browse%20by%20Day&fa=&fb=&fc=&fd=&all=1"

OUTPUT_DIR = Path("output")
CACHE_DIR = Path("cache")
CACHE_SESSIONS_DIR = CACHE_DIR / "sessions"
CACHE_PRESENTATIONS_DIR = CACHE_DIR / "presentations"

for d in [OUTPUT_DIR, CACHE_DIR, CACHE_SESSIONS_DIR, CACHE_PRESENTATIONS_DIR]:
    d.mkdir(parents=True, exist_ok=True)

MASTER_XLSX = OUTPUT_DIR / "acr_2026_programme.xlsx"
MASTER_JSON = OUTPUT_DIR / "acr_2026_programme.json"

MAX_WORKERS = 10
REQUEST_TIMEOUT = 25
MAX_RETRIES = 5
RETRY_BACKOFF = 1.0

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "X-Requested-With": "XMLHttpRequest",
}

EXCEL_HEADERS = [
    "Session_ID", "Session_Title", "Session_Type", "Room", "Date", "Day", "Time", "Chairs", "Chairs_Geography", "Session_Description", "Session_URL",
    "Presentation_ID", "Presentation_Title", "Presenter", "Presentation_Time", "Authors", "Authors_and_Affiliations", "Presenter_Geography", "Presentation_Type", "Keywords", "Abstract_Full_Text", "Acknowledgments_and_Disclosures", "Presentation_URL"
]

WRAP_COLUMNS = {"Session_Description", "Authors_and_Affiliations", "Abstract_Full_Text", "Acknowledgments_and_Disclosures"}

# ---------------------------------------------------------------------------
# Helper Functions
# ---------------------------------------------------------------------------

def clean_text(text: Optional[str]) -> str:
    """Clean HTML tags, unescape entities, strip replacement characters, and normalize whitespace."""
    if not text:
        return ""
    t = _html.unescape(str(text))
    t = re.sub(r"<[a-zA-Z/][^>]*>", " ", t)
    t = t.replace("\xa0", " ").replace("\ufffd", "").replace("\xc2", "").replace(chr(194), "").replace("​", "").replace("­", "")
    t = re.sub(r"[ \t]+", " ", t)
    t = re.sub(r"\n\s*\n+", "\n\n", t)
    return t.strip(" \t\n\r\ufffd\xc2" + chr(194))


def strip_control_chars(val: Any) -> Any:
    """Clean control characters incompatible with Excel XML."""
    if isinstance(val, str):
        return ILLEGAL_CHARACTERS_RE.sub("", val)
    return val


def parse_date_and_day(date_raw: str) -> tuple[str, str]:
    """Parse date string into (Date, Day) e.g., 'Friday, November 6, 2026' -> ('06 November 2026', 'Friday')."""
    if not date_raw:
        return "", ""
    date_clean = clean_text(date_raw)
    try:
        dt = datetime.strptime(date_clean, "%A, %B %d, %Y")
        return dt.strftime("%d %B %Y"), dt.strftime("%A")
    except Exception:
        pass
    parts = [p.strip() for p in date_clean.split(",") if p.strip()]
    if len(parts) >= 2:
        return ", ".join(parts[1:]), parts[0]
    return date_clean, ""

# ---------------------------------------------------------------------------
# Data Classes
# ---------------------------------------------------------------------------

@dataclass
class PresenterInfo:
    presenter_id: str = ""
    name: str = ""
    role: str = ""
    title_and_dept: str = ""
    institution: str = ""
    location: str = ""  # City, State/Country
    full_affiliation: str = ""


@dataclass
class PresentationData:
    presentation_id: str = ""
    presentation_code: str = ""  # buildcode
    title: str = ""
    date: str = ""
    day: str = ""
    time_range: str = ""
    presentation_type: str = ""
    keywords: str = ""
    presenters: list[PresenterInfo] = field(default_factory=list)
    presenter_names: str = ""
    presenter_geography: str = ""
    authors_and_affiliations: str = ""
    abstract_full_text: str = ""
    disclosures: str = ""
    presentation_url: str = ""


@dataclass
class SessionData:
    session_id: str = ""
    title: str = ""
    session_type: str = ""
    room: str = ""
    date: str = ""
    day: str = ""
    time_range: str = ""
    chairs: list[PresenterInfo] = field(default_factory=list)
    chairs_names: str = ""
    chairs_geography: str = ""
    description: str = ""
    session_url: str = ""
    presentations: list[PresentationData] = field(default_factory=list)

# ---------------------------------------------------------------------------
# HTTP & Cache Helpers
# ---------------------------------------------------------------------------

def fetch_url(url: str, is_ajax: bool = True, cache_path: Optional[Path] = None, force: bool = False) -> str:
    """Fetch URL with disk caching and auto retries."""
    if cache_path and cache_path.exists() and not force:
        try:
            return cache_path.read_text(encoding="utf-8")
        except Exception:
            pass

    req_headers = HEADERS if is_ajax else {"User-Agent": HEADERS["User-Agent"]}
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            r = requests.get(url, headers=req_headers, timeout=REQUEST_TIMEOUT)
            if r.status_code == 200:
                html_content = r.text
                if cache_path:
                    cache_path.write_text(html_content, encoding="utf-8")
                return html_content
            log.warning("HTTP %s on %s (attempt %d/%d)", r.status_code, url, attempt, MAX_RETRIES)
            time.sleep(RETRY_BACKOFF * attempt)
        except Exception as e:
            log.warning("Fetch error on %s (attempt %d): %s", url, attempt, e)
            time.sleep(RETRY_BACKOFF * attempt)
    return ""

# ---------------------------------------------------------------------------
# Parsing Logic
# ---------------------------------------------------------------------------

def parse_speakers(soup: BeautifulSoup) -> list[PresenterInfo]:
    """Parse speakers/faculty/chairs from popup HTML (`ul.speakers-wrap`)."""
    speakers = []
    speakers_wrap = soup.find("ul", class_="speakers-wrap")
    if not speakers_wrap:
        return speakers

    current_role = "Speaker"
    for child in speakers_wrap.children:
        if child.name == "h2" and "role-title" in child.get("class", []):
            role_text = clean_text(child.get_text())
            current_role = re.sub(r"\(s\)$", "", role_text, flags=re.I).strip()
        elif child.name == "li" and "speakerrow" in child.get("class", []):
            pid = child.get("data-presenterid", "")
            name_p = child.find("p", class_="speaker-name")
            name = clean_text(name_p.get_text()) if name_p else ""
            if not name:
                name_link = child.find("a", class_="loadbyurl")
                if name_link:
                    name = clean_text(name_link.get_text())

            prof_p = child.find("p", class_="prof-text")
            inst, loc = "", ""
            if prof_p:
                lines = [clean_text(line) for line in prof_p.get_text(separator="\n").split("\n") if clean_text(line)]
                if len(lines) >= 1:
                    inst = lines[0]
                if len(lines) >= 2:
                    loc = lines[1]

            full_aff = f"{inst}, {loc}".strip(", ") if (inst or loc) else ""
            speakers.append(PresenterInfo(
                presenter_id=pid,
                name=name,
                role=current_role,
                institution=inst,
                location=loc,
                full_affiliation=full_aff
            ))

    return speakers


def parse_session_popup(html: str, session_id: str) -> SessionData:
    """Parse SessionInfo AJAX popup HTML into SessionData."""
    sd = SessionData(session_id=session_id)
    sd.session_url = f"{BASE_URL}/ajaxcalls/SessionInfo.asp?PresentationID={session_id}"

    if not html:
        return sd

    soup = BeautifulSoup(html, "html.parser")
    popup_content = soup.find("div", class_="popup_content") or soup

    # Title
    h1 = popup_content.find("h1")
    if h1:
        sd.title = clean_text(h1.get_text())

    # Date & Time
    tidbits = popup_content.find_all("div", class_="pres-tidbit")
    for tb in tidbits:
        text = clean_text(tb.get_text())
        if tb.find("i", class_="fa-calendar"):
            sd.date, sd.day = parse_date_and_day(text)
        elif tb.find("i", class_="fa-clock-o"):
            sd.time_range = text

    # Tracks / General Topic
    tracks = [clean_text(t.get_text()) for t in popup_content.find_all("p", class_="trackname")]
    topic_div = popup_content.find(lambda el: el.name == "div" and "General Topic:" in el.text)
    if topic_div:
        tracks.append(clean_text(topic_div.get_text().replace("General Topic:", "")))
    sd.session_type = " | ".join(dict.fromkeys(t for t in tracks if t))

    # Room / Location
    room_div = popup_content.find("div", class_="roomname") or popup_content.find("span", class_="room")
    if room_div:
        sd.room = clean_text(room_div.get_text())

    # Chairs / Faculty
    sd.chairs = parse_speakers(soup)
    sd.chairs_names = " | ".join(f"{c.name} ({c.role})" if c.role else c.name for c in sd.chairs if c.name)
    sd.chairs_geography = " | ".join(c.full_affiliation for c in sd.chairs if c.full_affiliation)

    # Session Description / Learning Objectives
    desc_parts = []
    desc_div = popup_content.find("div", class_="presentation-description") or popup_content.find("div", id="session-description")
    if desc_div:
        desc_parts.append(clean_text(desc_div.get_text()))

    # Find paragraphs following headers
    for sec_title in popup_content.find_all(["h2", "h3", "h4", "strong"]):
        st_text = clean_text(sec_title.get_text())
        if any(keyword in st_text.lower() for keyword in ["objective", "learning", "overview", "description"]):
            nxt = sec_title.find_next_sibling()
            if nxt:
                desc_parts.append(f"{st_text}:\n{clean_text(nxt.get_text())}")

    sd.description = "\n\n".join(dict.fromkeys(p for p in desc_parts if p))

    return sd


def parse_presentation_popup(html: str, pres_id: str) -> PresentationData:
    """Parse PresentationInfo AJAX popup HTML into PresentationData."""
    pd_obj = PresentationData(presentation_id=pres_id)
    pd_obj.presentation_url = f"{BASE_URL}/ajaxcalls/PresentationInfo.asp?PresentationID={pres_id}"

    if not html:
        return pd_obj

    soup = BeautifulSoup(html, "html.parser")
    popup_content = soup.find("div", class_="popup_content") or soup

    # Title
    h1 = popup_content.find("h1")
    if h1:
        pd_obj.title = clean_text(h1.get_text())

    # Date & Time
    tidbits = popup_content.find_all("div", class_="pres-tidbit")
    for tb in tidbits:
        text = clean_text(tb.get_text())
        if tb.find("i", class_="fa-calendar"):
            pd_obj.date, pd_obj.day = parse_date_and_day(text)
        elif tb.find("i", class_="fa-clock-o"):
            pd_obj.time_range = text

    # Tracks / Types
    tracks = [clean_text(t.get_text()) for t in popup_content.find_all("p", class_="trackname")]
    pd_obj.presentation_type = " | ".join(dict.fromkeys(t for t in tracks if t))

    # Presenters & Authors
    pd_obj.presenters = parse_speakers(soup)
    pd_obj.presenter_names = " | ".join(p.name for p in pd_obj.presenters if p.name)
    pd_obj.presenter_geography = " | ".join(p.full_affiliation for p in pd_obj.presenters if p.full_affiliation)

    # Format combined Authors_and_Affiliations
    if pd_obj.presenters:
        author_names = []
        affiliations = []
        aff_map = {}

        for p in pd_obj.presenters:
            if not p.name:
                continue
            if p.full_affiliation:
                if p.full_affiliation not in aff_map:
                    aff_map[p.full_affiliation] = len(aff_map) + 1
                idx = aff_map[p.full_affiliation]
                author_names.append(f"{p.name} ({idx})")
            else:
                author_names.append(p.name)

        for aff_text, idx in aff_map.items():
            affiliations.append(f"({idx}) {aff_text}")

        pd_obj.authors_and_affiliations = f"Authors: {', '.join(author_names)}\n\nAffiliations:\n" + "\n".join(affiliations)
    else:
        pd_obj.authors_and_affiliations = ""

    # Disclosures
    disc_blocks = [clean_text(d.get_text()) for d in popup_content.find_all("div", class_="presentation-disclosure-block")]
    pd_obj.disclosures = "\n".join(d for d in disc_blocks if d)

    # Abstract / Full Text / Content
    text_blocks = []
    # Check for abstract section elements
    for el in popup_content.find_all(["div", "section", "p"]):
        if el.get("class") and any(c in el.get("class") for c in ["abstract-text", "pres-description", "presentation-text"]):
            text_blocks.append(clean_text(el.get_text()))

    if not text_blocks:
        # Fallback to main content paragraphs ignoring metadata wrappers
        for p in popup_content.find_all("p"):
            if not p.get("class") or not any(c in p.get("class") for c in ["trackname", "speaker-name", "prof-text"]):
                t = clean_text(p.get_text())
                if t and len(t) > 30 and not t.startswith("Disclosure"):
                    text_blocks.append(t)

    pd_obj.abstract_full_text = "\n\n".join(dict.fromkeys(b for b in text_blocks if b))

    return pd_obj

# ---------------------------------------------------------------------------
# Agenda Discovery
# ---------------------------------------------------------------------------

def discover_agenda_structure(force: bool = False) -> list[SessionData]:
    """Scrape Full Schedule agenda page to discover all sessions and child presentations."""
    log.info("Fetching full schedule agenda from %s ...", AGENDA_ALL_URL)
    agenda_html = fetch_url(AGENDA_ALL_URL, is_ajax=False, cache_path=CACHE_DIR / "agenda_all.html", force=force)
    if not agenda_html:
        log.error("Failed to fetch agenda page.")
        return []

    soup = BeautifulSoup(agenda_html, "html.parser")
    agenda_ul = soup.find("ul", id="agenda")
    if not agenda_ul:
        log.error("Could not find #agenda list on page.")
        return []

    sessions: list[SessionData] = []
    current_date = ""
    current_day = ""

    children = list(agenda_ul.children)
    log.info("Processing %d agenda elements...", len(children))

    i = 0
    while i < len(children):
        elem = children[i]
        i += 1
        if not hasattr(elem, "name") or elem.name is None:
            continue

        # Date header row
        if elem.name == "li" and "dayrow" in elem.get("class", []):
            date_raw = clean_text(elem.get_text())
            current_date, current_day = parse_date_and_day(date_raw)
            continue

        # Session bucket row
        if elem.name == "li" and "bucket" in elem.get("class", []):
            session_title_span = elem.find("div", class_="list-row-primary")
            session_title = clean_text(session_title_span.get_text()) if session_title_span else ""

            time_span = elem.find("div", class_="prestime")
            time_range = clean_text(time_span.get_text()) if time_span else ""

            # Check next sibling div for bucketwrapper containing items
            bucket_wrapper = None
            if i < len(children) and children[i].name == "div" and "bucketwrapper" in children[i].get("class", []):
                bucket_wrapper = children[i]
                i += 1

            session_id = ""
            presentations: list[PresentationData] = []

            if bucket_wrapper:
                item_lis = bucket_wrapper.find_all("li", class_="loadbyurl")
                for item in item_lis:
                    pres_id = item.get("data-presid", "")
                    buildcode = item.get("data-buildcode", "")
                    url_attr = item.get("data-url", "")

                    if buildcode == "M" or "SessionInfo.asp" in url_attr:
                        session_id = pres_id
                    elif buildcode in ("P", "HP") or "PresentationInfo.asp" in url_attr:
                        p_title_span = item.find("div", class_="list-row-primary")
                        p_title = clean_text(p_title_span.get_text()) if p_title_span else ""
                        p_time_span = item.find("div", class_="prestime")
                        p_time = clean_text(p_time_span.get_text()) if p_time_span else ""

                        presentations.append(PresentationData(
                            presentation_id=pres_id,
                            presentation_code=buildcode,
                            title=p_title,
                            time_range=p_time,
                            date=current_date,
                            day=current_day
                        ))

            if not session_id and presentations:
                session_id = presentations[0].presentation_id

            if session_id or session_title:
                sessions.append(SessionData(
                    session_id=session_id,
                    title=session_title,
                    date=current_date,
                    day=current_day,
                    time_range=time_range,
                    presentations=presentations
                ))

    log.info("Discovered %d session buckets with total %d presentations.",
             len(sessions), sum(len(s.presentations) for s in sessions))
    return sessions

# ---------------------------------------------------------------------------
# Parallel Detail Extractor
# ---------------------------------------------------------------------------

def extract_all_details(sessions: list[SessionData], force: bool = False):
    """Concurrently fetch and parse SessionInfo and PresentationInfo details."""
    session_ids = set()
    presentation_ids = set()

    for s in sessions:
        if s.session_id:
            session_ids.add(s.session_id)
        for p in s.presentations:
            if p.presentation_id:
                presentation_ids.add(p.presentation_id)

    log.info("Starting detail extraction: %d unique sessions, %d unique presentations...",
             len(session_ids), len(presentation_ids))

    session_results: dict[str, SessionData] = {}
    presentation_results: dict[str, PresentationData] = {}

    def fetch_session_worker(sid: str) -> tuple[str, SessionData]:
        url = f"{BASE_URL}/ajaxcalls/SessionInfo.asp?PresentationID={sid}"
        cache_file = CACHE_SESSIONS_DIR / f"{sid}.html"
        html = fetch_url(url, is_ajax=True, cache_path=cache_file, force=force)
        return sid, parse_session_popup(html, sid)

    def fetch_presentation_worker(pid: str) -> tuple[str, PresentationData]:
        url = f"{BASE_URL}/ajaxcalls/PresentationInfo.asp?PresentationID={pid}"
        cache_file = CACHE_PRESENTATIONS_DIR / f"{pid}.html"
        html = fetch_url(url, is_ajax=True, cache_path=cache_file, force=force)
        return pid, parse_presentation_popup(html, pid)

    # 1. Fetch Session details
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futures = {executor.submit(fetch_session_worker, sid): sid for sid in session_ids}
        for future in as_completed(futures):
            sid, sd_parsed = future.result()
            session_results[sid] = sd_parsed

    # 2. Fetch Presentation details
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futures = {executor.submit(fetch_presentation_worker, pid): pid for pid in presentation_ids}
        for future in as_completed(futures):
            pid, pd_parsed = future.result()
            presentation_results[pid] = pd_parsed

    # Merge detail results back into sessions data model
    for s in sessions:
        if s.session_id in session_results:
            sd_parsed = session_results[s.session_id]
            if sd_parsed.title:
                s.title = sd_parsed.title
            if sd_parsed.session_type:
                s.session_type = sd_parsed.session_type
            if sd_parsed.room:
                s.room = sd_parsed.room
            if sd_parsed.time_range:
                s.time_range = sd_parsed.time_range
            if sd_parsed.chairs:
                s.chairs = sd_parsed.chairs
                s.chairs_names = sd_parsed.chairs_names
                s.chairs_geography = sd_parsed.chairs_geography
            if sd_parsed.description:
                s.description = sd_parsed.description
            s.session_url = sd_parsed.session_url

        for i, p in enumerate(s.presentations):
            if p.presentation_id in presentation_results:
                pd_parsed = presentation_results[p.presentation_id]
                if pd_parsed.title:
                    p.title = pd_parsed.title
                if pd_parsed.time_range:
                    p.time_range = pd_parsed.time_range
                if pd_parsed.presentation_type:
                    p.presentation_type = pd_parsed.presentation_type
                if pd_parsed.presenters:
                    p.presenters = pd_parsed.presenters
                    p.presenter_names = pd_parsed.presenter_names
                    p.presenter_geography = pd_parsed.presenter_geography
                if pd_parsed.authors_and_affiliations:
                    p.authors_and_affiliations = pd_parsed.authors_and_affiliations
                if pd_parsed.abstract_full_text:
                    p.abstract_full_text = pd_parsed.abstract_full_text
                if pd_parsed.disclosures:
                    p.disclosures = pd_parsed.disclosures
                p.presentation_url = pd_parsed.presentation_url

# ---------------------------------------------------------------------------
# Excel Workbook Generation
# ---------------------------------------------------------------------------

def write_excel_workbook(sessions: list[SessionData], output_path: Path):
    """Write master workbook using OpenPyXL with professional styling."""
    log.info("Generating master Excel workbook at %s ...", output_path)
    wb = Workbook()
    ws = wb.active
    ws.title = "Presentations"

    # Header Row
    ws.append(EXCEL_HEADERS)
    header_font = Font(name="Segoe UI", size=11, bold=True, color="000000")
    header_fill = PatternFill(start_color="D9E9F8", end_color="D9E9F8", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
    ws.freeze_panes = "A2"

    wrap_align = Alignment(vertical="top", wrap_text=True)
    std_align = Alignment(vertical="top", wrap_text=False)
    wrap_col_indices = {EXCEL_HEADERS.index(h) + 1 for h in WRAP_COLUMNS}

    total_rows = 0
    for s in sessions:
        rows_to_write = []
        if s.presentations:
            for p in s.presentations:
                rows_to_write.append({
                    "Session_ID": s.session_id,
                    "Session_Title": s.title,
                    "Session_Type": s.session_type,
                    "Room": s.room,
                    "Date": s.date,
                    "Day": s.day,
                    "Time": s.time_range,
                    "Chairs": s.chairs_names,
                    "Chairs_Geography": s.chairs_geography,
                    "Session_Description": s.description,
                    "Session_URL": s.session_url,
                    "Presentation_ID": p.presentation_id,
                    "Presentation_Title": p.title,
                    "Presenter": p.presenter_names,
                    "Presentation_Time": p.time_range,
                    "Authors": p.presenter_names,
                    "Authors_and_Affiliations": p.authors_and_affiliations,
                    "Presenter_Geography": p.presenter_geography,
                    "Presentation_Type": p.presentation_type,
                    "Keywords": p.keywords,
                    "Abstract_Full_Text": p.abstract_full_text,
                    "Acknowledgments_and_Disclosures": p.disclosures,
                    "Presentation_URL": p.presentation_url
                })
        else:
            # Session with 0 presentations
            rows_to_write.append({
                "Session_ID": s.session_id,
                "Session_Title": s.title,
                "Session_Type": s.session_type,
                "Room": s.room,
                "Date": s.date,
                "Day": s.day,
                "Time": s.time_range,
                "Chairs": s.chairs_names,
                "Chairs_Geography": s.chairs_geography,
                "Session_Description": s.description,
                "Session_URL": s.session_url,
                "Presentation_ID": "",
                "Presentation_Title": "",
                "Presenter": "",
                "Presentation_Time": "",
                "Authors": "",
                "Authors_and_Affiliations": "",
                "Presenter_Geography": "",
                "Presentation_Type": "",
                "Keywords": "",
                "Abstract_Full_Text": "",
                "Acknowledgments_and_Disclosures": "",
                "Presentation_URL": ""
            })

        for r_dict in rows_to_write:
            row_vals = [strip_control_chars(r_dict.get(h, "")) for h in EXCEL_HEADERS]
            ws.append(row_vals)
            curr_row = ws.max_row
            total_rows += 1

            for col_idx in range(1, len(EXCEL_HEADERS) + 1):
                cell = ws.cell(row=curr_row, column=col_idx)
                if col_idx in wrap_col_indices:
                    cell.alignment = wrap_align
                else:
                    cell.alignment = std_align

    # Column Dimensions
    for idx, header in enumerate(EXCEL_HEADERS, start=1):
        col_letter = get_column_letter(idx)
        if header in WRAP_COLUMNS:
            ws.column_dimensions[col_letter].width = 55
        elif header in ("Session_Title", "Presentation_Title"):
            ws.column_dimensions[col_letter].width = 40
        else:
            ws.column_dimensions[col_letter].width = 22

    wb.save(output_path)
    log.info("Workbook saved successfully with %d total data rows.", total_rows)

# ---------------------------------------------------------------------------
# Main Routine
# ---------------------------------------------------------------------------

def main():
    import argparse
    parser = argparse.ArgumentParser(description="ACR Convergence 2026 Scraper")
    parser.add_argument("--force", "--refresh", action="store_true", help="Force refresh by bypassing local disk cache")
    args = parser.parse_args()

    if args.force:
        log.info("Starting ACR Convergence 2026 Scraper (FORCE REFRESH ENABLED - BYPASSING DISK CACHE)...")
    else:
        log.info("Starting ACR Convergence 2026 Scraper (Using local disk cache if available)...")

    start_time = time.time()

    # 1. Discover Agenda Structure
    sessions = discover_agenda_structure(force=args.force)
    if not sessions:
        log.error("No sessions discovered. Exiting.")
        sys.exit(1)

    # 2. Extract Details
    extract_all_details(sessions, force=args.force)

    # 3. Save JSON Master Backup
    json_data = [
        {
            "session_id": s.session_id,
            "title": s.title,
            "session_type": s.session_type,
            "room": s.room,
            "date": s.date,
            "day": s.day,
            "time_range": s.time_range,
            "chairs_names": s.chairs_names,
            "chairs_geography": s.chairs_geography,
            "description": s.description,
            "session_url": s.session_url,
            "presentations": [
                {
                    "presentation_id": p.presentation_id,
                    "title": p.title,
                    "time_range": p.time_range,
                    "presenter_names": p.presenter_names,
                    "authors_and_affiliations": p.authors_and_affiliations,
                    "presenter_geography": p.presenter_geography,
                    "presentation_type": p.presentation_type,
                    "abstract_full_text": p.abstract_full_text,
                    "disclosures": p.disclosures,
                    "presentation_url": p.presentation_url
                }
                for p in s.presentations
            ]
        }
        for s in sessions
    ]
    with open(MASTER_JSON, "w", encoding="utf-8") as f:
        json.dump(json_data, f, ensure_ascii=False, indent=2)
    log.info("Saved raw JSON backup at %s", MASTER_JSON)

    # 4. Generate Master Excel Output
    write_excel_workbook(sessions, MASTER_XLSX)

    elapsed = time.time() - start_time
    log.info("Scraping completed in %.2f seconds.", elapsed)


if __name__ == "__main__":
    main()
